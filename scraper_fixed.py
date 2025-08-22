import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
import time
import random
import logging
import os
from urllib.parse import urlparse, urlunparse
import re

# Configure logging
logging.basicConfig(filename='debug.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# List of User-Agent strings
user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 11_5_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36 Edg/94.0.992.50",
]

class CompanyWebsiteScraper:
    def __init__(self):
        # No CSV data needed for fully automated web search
        pass
    
    # No CSV methods needed for fully automated web search

def is_valid_url(url):
    """Check if the given string is a valid URL"""
    regex = re.compile(
        r'^(?:http|ftp)s?://'  
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|' 
        r'localhost|'  # Localhost
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}|'  # IPv4
        r'\[?[A-F0-9]*:[A-F0-9:]+\]?)'  # IPv6
        r'(?::\d+)?'  # Port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)
    return re.match(regex, url) is not None

def clean_url(url):
    """Clean and normalize URL"""
    parsed_url = urlparse(url)
    cleaned_url = urlunparse((parsed_url.scheme, parsed_url.netloc, '', '', '', ''))
    return cleaned_url

def is_general_website(url):
    """Check if URL is from general/non-company websites to avoid"""
    if not url:
        return True
    
    # Convert to lowercase for case-insensitive matching
    url_lower = url.lower()
    
    # List of general websites to avoid
    general_domains = [
        # Social Media
        'facebook.com', 'twitter.com', 'linkedin.com', 'instagram.com', 
        'youtube.com', 'tiktok.com', 'snapchat.com', 'pinterest.com',
        
        # Wikipedia and Reference
        'wikipedia.org', 'wikimedia.org', 'wikidata.org',
        
        # General Business Directories
        'yellowpages.com', 'yelp.com', 'foursquare.com', 'google.com',
        'bing.com', 'yahoo.com', 'duckduckgo.com', 'startpage.com',
        
        # News and Media
        'cnn.com', 'bbc.com', 'reuters.com', 'bloomberg.com', 'forbes.com',
        'wsj.com', 'nytimes.com', 'washingtonpost.com',
        
        # General Platforms
        'amazon.com', 'ebay.com', 'alibaba.com', 'etsy.com',
        'craigslist.org', 'gumtree.com', 'indeed.com', 'glassdoor.com',
        
        # File Sharing and Cloud
        'dropbox.com', 'drive.google.com', 'onedrive.com', 'box.com',
        
        # Forums and Q&A
        'reddit.com', 'quora.com', 'stackoverflow.com', 'stackexchange.com',
        
        # Government and Educational (too general)
        '.gov', '.edu', '.ac.uk', '.edu.au',
        
        # Business Profile Sites
        'crunchbase.com', 'owler.com', 'zoominfo.com', 'dnb.com',
        'manta.com', 'spoke.com', 'businesswire.com', 'prnewswire.com'
    ]
    
    # Check if URL contains any general domains
    for domain in general_domains:
        if domain in url_lower:
            return True
    
    return False

def get_first_three_words(company_name):
    """Extract first four words from company name for search"""
    if not company_name:
        return ""
    
    # Clean up the company name and split into words
    words = str(company_name).strip().split()
    
    # Take only the first 4 words
    first_four = words[:4]
    
    # Join them back together
    return ' '.join(first_four)

def get_company_website_enhanced(company_name, scraper_instance):
    """Fully automated website search using web search engines only"""
    try:
        # Use web search with first 4 words only for full automation
        search_name = get_first_three_words(company_name)
        logging.info(f"Searching with first 4 words: '{company_name}' -> '{search_name}'")
        
        search_engines = [
            ("https://duckduckgo.com/?q=", 'a[data-testid="result-title-a"]'),
            ("https://www.startpage.com/do/search?q=", 'a.w-gl__result-title'),
        ]
        
        for search_url_base, selector in search_engines:
            try:
                # Configure Chrome options for this specific search
                chrome_options = Options()
                chrome_options.add_argument("--headless")
                chrome_options.add_argument("--disable-gpu")
                chrome_options.add_argument("--no-sandbox")
                chrome_options.add_argument("--disable-dev-shm-usage")
                chrome_options.add_argument(f"user-agent={random.choice(user_agents)}")
                chrome_options.add_argument("--disable-blink-features=AutomationControlled")
                chrome_options.add_argument("--disable-infobars")
                
                # Create a fresh driver for each search to avoid detection
                service = Service(ChromeDriverManager().install())
                driver = webdriver.Chrome(service=service, options=chrome_options)
                
                try:
                    # Perform the search using first 3 words only
                    search_query = search_name.replace(' ', '+')
                    search_url = f"{search_url_base}{search_query}"
                    driver.get(search_url)

                    # Wait for the page to load
                    time.sleep(random.uniform(5, 8))

                    # Simulate human-like behavior occasionally
                    if random.random() < 0.3:
                        actions = ActionChains(driver)
                        actions.move_by_offset(random.randint(10, 50), random.randint(10, 50)).perform()
                        time.sleep(random.uniform(1, 2))

                    # Check for blocked content
                    page_source = driver.page_source
                    if any(word in page_source.lower() for word in ["captcha", "sorry", "blocked", "unusual traffic"]):
                        logging.warning(f"Search blocked for {company_name} on {search_url_base}")
                        continue

                    # Extract search result links and filter out general websites
                    try:
                        # Try to find multiple results to filter through
                        results = driver.find_elements(By.CSS_SELECTOR, selector)
                        if not results:
                            # Try alternative selector for DuckDuckGo
                            if "duckduckgo" in search_url_base:
                                results = driver.find_elements(By.CSS_SELECTOR, 'h2 a')
                        
                        # Check each result to find a valid company website
                        for result in results[:5]:  # Check first 5 results
                            try:
                                website = result.get_attribute('href')
                                
                                if website and is_valid_url(website):
                                    # Skip general websites
                                    if is_general_website(website):
                                        logging.debug(f"Skipping general website: {website}")
                                        continue
                                    
                                    # Clean up the URL
                                    website = clean_url(website)
                                    logging.info(f"Found valid company website: {company_name} -> {website}")
                                    return website
                            except Exception as e:
                                logging.debug(f"Error processing result: {e}")
                                continue
                        
                        # If no valid results found with primary selector, log it
                        logging.debug(f"No valid company websites found in results for {company_name}")
                        continue
                            
                    except Exception as e:
                        logging.debug(f"No results found with selector {selector} for {company_name}: {e}")
                        continue
                        
                finally:
                    driver.quit()
                    
            except Exception as e:
                logging.error(f"Search engine error for {company_name} on {search_url_base}: {e}")
                continue

        # If all search engines fail, return failure
        logging.error(f"All search methods failed for {company_name}")
        return "Search failed after retries"
        
    except Exception as e:
        logging.error(f"Critical error for {company_name}: {e}")
        return "Request failed"

def process_company(row, sheet, scraper_instance):
    """Process a single company row"""
    company_name = sheet.cell(row=row, column=2).value
    website_cell = sheet.cell(row=row, column=3)

    # Skip if the website cell already contains a valid URL
    if website_cell.value and is_valid_url(website_cell.value):
        return

    if company_name:
        website = get_company_website_enhanced(company_name, scraper_instance)
        website_cell.value = website
        logging.info(f"Processed row {row}: {company_name} -> {website}")
        print(f"Processed: {company_name}, Website: {website}")
        
        # Add a delay between requests to be respectful
        time.sleep(random.uniform(15, 25))

def main():
    try:
        # Simple validation without CSV dependency
        print("🚀 Starting Fully Automated Company Website Scraper")
        print("=" * 50)
        
        excel_file = 'Company_Website.xlsx'
        
        # Load Excel file to check company count first
        wb = openpyxl.load_workbook(excel_file)
        sheet_name = '1. Exhibitor List (Input)'
        sheet = wb[sheet_name]

        # Count total companies in the Excel file
        total_companies = 0
        for row in range(9, sheet.max_row + 1):
            company_name = sheet.cell(row=row, column=2).value
            if company_name and str(company_name).strip():
                total_companies += 1
        
        # Check if company count exceeds 300
        if total_companies > 300:
            error_msg = f"❌ ERROR: Too many companies detected ({total_companies})!"
            print(error_msg)
            print("🚫 This scraper accepts maximum 300 companies only.")
            print("💡 Please reduce the number of companies in your Excel file to 300 or fewer.")
            logging.error(f"Company limit exceeded: {total_companies} companies found, maximum allowed is 300")
            return
        
        print(f"✅ Company count check passed: {total_companies} companies (within 300 limit)")
        
        # Collect all rows that need processing
        rows_to_process = []
        for row in range(9, sheet.max_row + 1):
            website_cell = sheet.cell(row=row, column=3)
            if not website_cell.value or not is_valid_url(website_cell.value):
                rows_to_process.append(row)

        if not rows_to_process:
            print("✅ All companies already have valid websites!")
            return

        print(f"\n🔄 Processing {len(rows_to_process)} companies with web search only...")
        
        # Create scraper instance
        scraper = CompanyWebsiteScraper()
        
        # Process companies one by one (no threading to avoid blocking)
        for i, row in enumerate(rows_to_process, 1):
            print(f"\n📡 Processing {i}/{len(rows_to_process)}...")
            process_company(row, sheet, scraper)
            
            # Save progress every 5 companies
            if i % 5 == 0:
                wb.save(excel_file)
                print(f"💾 Progress saved ({i}/{len(rows_to_process)} completed)")

        # Save final results
        wb.save(excel_file)
        print("\n🎉 Excel file updated successfully!")
        print(f"📊 Processed {len(rows_to_process)} companies using web search only")
        
    except FileNotFoundError:
        logging.error("Excel file not found. Ensure the file exists.")
        print("❌ Excel file not found. Ensure the file exists.")
    except PermissionError:
        logging.error("Permission error: Ensure the file is not open elsewhere.")
        print("❌ Permission error: Ensure the file is not open elsewhere.")
    except Exception as e:
        logging.error(f"Unexpected error: {e}")
        print(f"❌ Unexpected error: {e}")

if __name__ == "__main__":
    main()
