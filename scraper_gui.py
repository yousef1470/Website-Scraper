import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import queue
import os
import sys
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
import time
import random
import logging
from urllib.parse import urlparse, urlunparse
import re

class ScraperGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("🚀 Company Website Scraper Pro")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')
        
        # Variables
        self.excel_file_path = tk.StringVar()
        self.total_companies = tk.StringVar(value="0")
        self.processed_companies = tk.StringVar(value="0")
        self.success_rate = tk.StringVar(value="0%")
        self.is_running = False
        self.scraper_thread = None
        self.log_queue = queue.Queue()
        
        # Setup logging for GUI
        self.setup_logging()
        
        # Create GUI elements
        self.create_widgets()
        
        # Start log monitoring
        self.check_log_queue()
        
    def setup_logging(self):
        """Setup logging to capture scraper output"""
        # Create a custom handler that sends logs to the queue
        class QueueHandler(logging.Handler):
            def __init__(self, log_queue):
                super().__init__()
                self.log_queue = log_queue
                
            def emit(self, record):
                self.log_queue.put(self.format(record))
        
        # Configure logging
        self.queue_handler = QueueHandler(self.log_queue)
        self.queue_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        # Add handler to root logger
        logging.getLogger().addHandler(self.queue_handler)
        logging.getLogger().setLevel(logging.INFO)
        
    def create_widgets(self):
        """Create all GUI widgets"""
        # Main title
        title_frame = tk.Frame(self.root, bg='#2c3e50', height=60)
        title_frame.pack(fill='x', padx=0, pady=0)
        title_frame.pack_propagate(False)
        
        title_label = tk.Label(title_frame, text="🚀 Company Website Scraper Pro", 
                              font=('Arial', 18, 'bold'), fg='white', bg='#2c3e50')
        title_label.pack(expand=True)
        
        # Main container
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # File selection section
        self.create_file_section(main_frame)
        
        # Statistics section
        self.create_stats_section(main_frame)
        
        # Control buttons section
        self.create_control_section(main_frame)
        
        # Progress section
        self.create_progress_section(main_frame)
        
        # Log section
        self.create_log_section(main_frame)
        
    def create_file_section(self, parent):
        """Create file selection section"""
        file_frame = tk.LabelFrame(parent, text="📁 Excel File Selection", 
                                  font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        file_frame.pack(fill='x', pady=(0, 15))
        
        # File path entry
        path_frame = tk.Frame(file_frame, bg='#f0f0f0')
        path_frame.pack(fill='x', padx=15, pady=15)
        
        tk.Label(path_frame, text="Excel File:", font=('Arial', 10), bg='#f0f0f0').pack(anchor='w')
        
        entry_frame = tk.Frame(path_frame, bg='#f0f0f0')
        entry_frame.pack(fill='x', pady=(5, 0))
        
        self.file_entry = tk.Entry(entry_frame, textvariable=self.excel_file_path, 
                                  font=('Arial', 10), state='readonly', bg='white')
        self.file_entry.pack(side='left', fill='x', expand=True)
        
        browse_btn = tk.Button(entry_frame, text="📁 Browse", command=self.browse_file,
                              bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                              relief='flat', padx=20)
        browse_btn.pack(side='right', padx=(10, 0))
        
        # File info
        self.file_info_label = tk.Label(path_frame, text="No file selected", 
                                       font=('Arial', 9), fg='#7f8c8d', bg='#f0f0f0')
        self.file_info_label.pack(anchor='w', pady=(5, 0))
        
    def create_stats_section(self, parent):
        """Create statistics section"""
        stats_frame = tk.LabelFrame(parent, text="📊 Statistics", 
                                   font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        stats_frame.pack(fill='x', pady=(0, 15))
        
        stats_container = tk.Frame(stats_frame, bg='#f0f0f0')
        stats_container.pack(fill='x', padx=15, pady=15)
        
        # Create stat boxes
        stat_boxes = [
            ("Total Companies", self.total_companies, "#3498db"),
            ("Processed", self.processed_companies, "#2ecc71"),
            ("Success Rate", self.success_rate, "#e74c3c")
        ]
        
        for i, (label, var, color) in enumerate(stat_boxes):
            box = tk.Frame(stats_container, bg=color, relief='raised', bd=2)
            box.pack(side='left', fill='both', expand=True, padx=(0, 10) if i < 2 else (0, 0))
            
            tk.Label(box, text=label, font=('Arial', 10), bg=color, fg='white').pack(pady=(10, 0))
            tk.Label(box, textvariable=var, font=('Arial', 16, 'bold'), 
                    bg=color, fg='white').pack(pady=(0, 10))
    
    def create_control_section(self, parent):
        """Create control buttons section"""
        control_frame = tk.LabelFrame(parent, text="🎮 Controls", 
                                     font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        control_frame.pack(fill='x', pady=(0, 15))
        
        button_frame = tk.Frame(control_frame, bg='#f0f0f0')
        button_frame.pack(pady=15)
        
        # Start button
        self.start_btn = tk.Button(button_frame, text="▶️ Start Scraping", 
                                  command=self.start_scraping,
                                  bg='#2ecc71', fg='white', font=('Arial', 12, 'bold'),
                                  relief='flat', padx=30, pady=10)
        self.start_btn.pack(side='left', padx=(0, 10))
        
        # Stop button
        self.stop_btn = tk.Button(button_frame, text="⏹️ Stop", 
                                 command=self.stop_scraping,
                                 bg='#e74c3c', fg='white', font=('Arial', 12, 'bold'),
                                 relief='flat', padx=30, pady=10, state='disabled')
        self.stop_btn.pack(side='left', padx=(0, 10))
        
        # Validate button
        self.validate_btn = tk.Button(button_frame, text="✅ Validate", 
                                     command=self.validate_file,
                                     bg='#f39c12', fg='white', font=('Arial', 12, 'bold'),
                                     relief='flat', padx=30, pady=10)
        self.validate_btn.pack(side='left')
        
    def create_progress_section(self, parent):
        """Create progress section"""
        progress_frame = tk.LabelFrame(parent, text="📈 Progress", 
                                      font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        progress_frame.pack(fill='x', pady=(0, 15))
        
        progress_container = tk.Frame(progress_frame, bg='#f0f0f0')
        progress_container.pack(fill='x', padx=15, pady=15)
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_container, variable=self.progress_var,
                                           maximum=100, length=400, mode='determinate')
        self.progress_bar.pack(fill='x', pady=(0, 10))
        
        # Progress label
        self.progress_label = tk.Label(progress_container, text="Ready to start", 
                                      font=('Arial', 10), bg='#f0f0f0', fg='#2c3e50')
        self.progress_label.pack()
        
    def create_log_section(self, parent):
        """Create log section"""
        log_frame = tk.LabelFrame(parent, text="📝 Activity Log", 
                                 font=('Arial', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        log_frame.pack(fill='both', expand=True)
        
        # Log text area
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, 
                                                 font=('Consolas', 9), bg='#2c3e50', fg='#ecf0f1',
                                                 insertbackground='white')
        self.log_text.pack(fill='both', expand=True, padx=15, pady=15)
        
        # Add welcome message
        self.log_message("🚀 Company Website Scraper Pro - Ready")
        self.log_message("📋 Features:")
        self.log_message("   • Supports up to 300 companies")
        self.log_message("   • Uses first 4 words for search")
        self.log_message("   • Filters out general websites")
        self.log_message("   • Multiple search engines (DuckDuckGo, Startpage)")
        self.log_message("   • Automatic progress saving")
        self.log_message("=" * 60)
        
    def browse_file(self):
        """Browse for Excel file"""
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filename:
            self.excel_file_path.set(filename)
            self.validate_file()
            
    def validate_file(self):
        """Validate the selected Excel file"""
        file_path = self.excel_file_path.get()
        
        if not file_path:
            self.file_info_label.config(text="❌ No file selected", fg='#e74c3c')
            return False
            
        if not os.path.exists(file_path):
            self.file_info_label.config(text="❌ File not found", fg='#e74c3c')
            return False
            
        try:
            # Load Excel file
            wb = openpyxl.load_workbook(file_path)
            
            # Check if required sheet exists
            if '1. Exhibitor List (Input)' not in wb.sheetnames:
                self.file_info_label.config(text="❌ Sheet '1. Exhibitor List (Input)' not found", fg='#e74c3c')
                return False
                
            sheet = wb['1. Exhibitor List (Input)']
            
            # Count companies
            company_count = 0
            for row in range(9, sheet.max_row + 1):
                company_name = sheet.cell(row=row, column=2).value
                if company_name and str(company_name).strip():
                    company_count += 1
                    
            self.total_companies.set(str(company_count))
            
            # Check company limit
            if company_count > 300:
                self.file_info_label.config(text=f"❌ Too many companies ({company_count}). Max 300 allowed.", fg='#e74c3c')
                self.log_message(f"❌ ERROR: {company_count} companies found. Maximum 300 allowed.")
                return False
                
            # File is valid
            self.file_info_label.config(text=f"✅ Valid file with {company_count} companies", fg='#2ecc71')
            self.log_message(f"✅ File validated: {company_count} companies found")
            return True
            
        except Exception as e:
            self.file_info_label.config(text=f"❌ Error reading file: {str(e)}", fg='#e74c3c')
            self.log_message(f"❌ Error validating file: {str(e)}")
            return False
            
    def start_scraping(self):
        """Start the scraping process"""
        if not self.validate_file():
            messagebox.showerror("Error", "Please select a valid Excel file first.")
            return
            
        if self.is_running:
            return
            
        self.is_running = True
        self.start_btn.config(state='disabled')
        self.stop_btn.config(state='normal')
        self.validate_btn.config(state='disabled')
        
        # Reset progress
        self.progress_var.set(0)
        self.processed_companies.set("0")
        self.success_rate.set("0%")
        
        # Start scraping in separate thread
        self.scraper_thread = threading.Thread(target=self.run_scraper, daemon=True)
        self.scraper_thread.start()
        
        self.log_message("🚀 Starting scraping process...")
        
    def stop_scraping(self):
        """Stop the scraping process"""
        self.is_running = False
        self.start_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.validate_btn.config(state='normal')
        
        self.progress_label.config(text="❌ Stopped by user")
        self.log_message("⏹️ Scraping stopped by user")
        
    def run_scraper(self):
        """Run the scraper in background thread"""
        try:
            # Import scraper functions
            from scraper_fixed import (
                CompanyWebsiteScraper, 
                get_company_website_enhanced, 
                is_valid_url, 
                process_company
            )
            
            excel_file = self.excel_file_path.get()
            
            # Load Excel file
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb['1. Exhibitor List (Input)']
            
            # Collect rows to process
            rows_to_process = []
            for row in range(9, sheet.max_row + 1):
                if not self.is_running:
                    return
                    
                website_cell = sheet.cell(row=row, column=3)
                if not website_cell.value or not is_valid_url(website_cell.value):
                    rows_to_process.append(row)
                    
            if not rows_to_process:
                self.log_message("✅ All companies already have valid websites!")
                self.progress_label.config(text="✅ All companies processed")
                self.stop_scraping()
                return
                
            total_to_process = len(rows_to_process)
            self.log_message(f"🔄 Processing {total_to_process} companies...")
            
            # Create scraper instance
            scraper = CompanyWebsiteScraper()
            
            successful = 0
            
            # Process companies
            for i, row in enumerate(rows_to_process, 1):
                if not self.is_running:
                    return
                    
                company_name = sheet.cell(row=row, column=2).value
                self.progress_label.config(text=f"Processing: {company_name}")
                self.log_message(f"📡 Processing {i}/{total_to_process}: {company_name}")
                
                # Get website
                website = get_company_website_enhanced(company_name, scraper)
                sheet.cell(row=row, column=3).value = website
                
                if is_valid_url(website):
                    successful += 1
                    self.log_message(f"✅ Found: {website}")
                else:
                    self.log_message(f"❌ Failed: {website}")
                
                # Update progress
                progress = (i / total_to_process) * 100
                self.progress_var.set(progress)
                self.processed_companies.set(str(i))
                self.success_rate.set(f"{(successful/i)*100:.1f}%")
                
                # Save progress every 5 companies
                if i % 5 == 0:
                    wb.save(excel_file)
                    self.log_message(f"💾 Progress saved ({i}/{total_to_process} completed)")
                    
                # Delay between requests
                if self.is_running:
                    time.sleep(random.uniform(15, 25))
                    
            # Final save
            wb.save(excel_file)
            self.log_message("🎉 Scraping completed successfully!")
            self.log_message(f"📊 Final results: {successful}/{total_to_process} successful")
            self.progress_label.config(text="🎉 Completed successfully!")
            
        except Exception as e:
            self.log_message(f"❌ Error during scraping: {str(e)}")
            self.progress_label.config(text="❌ Error occurred")
            
        finally:
            self.stop_scraping()
            
    def log_message(self, message):
        """Add message to log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        formatted_message = f"[{timestamp}] {message}\n"
        
        # Update in main thread
        self.root.after(0, self._update_log, formatted_message)
        
    def _update_log(self, message):
        """Update log in main thread"""
        self.log_text.insert(tk.END, message)
        self.log_text.see(tk.END)
        
    def check_log_queue(self):
        """Check for log messages from the queue"""
        try:
            while True:
                message = self.log_queue.get_nowait()
                self._update_log(f"{message}\n")
        except queue.Empty:
            pass
        finally:
            # Schedule next check
            self.root.after(100, self.check_log_queue)

def main():
    """Main function to run the GUI"""
    root = tk.Tk()
    app = ScraperGUI(root)
    
    # Set window icon (if available)
    try:
        root.iconbitmap('icon.ico')
    except:
        pass
        
    # Center window on screen
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

if __name__ == "__main__":
    main()
